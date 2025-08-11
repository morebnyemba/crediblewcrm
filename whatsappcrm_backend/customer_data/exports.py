# exports.py

import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count
from decimal import Decimal

# For PDF generation
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from .models import Payment, MemberProfile

# --- Helper Functions ---

def _auto_adjust_excel_columns(sheet):
    """
    Auto-adjusts the width of columns in an Excel sheet based on content length.
    """
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

def get_giver_name(payment):
    """Helper to get the best available name for a giver."""
    if payment.member and payment.member.get_full_name():
        return payment.member.get_full_name()
    if payment.contact and payment.contact.name:
        return payment.contact.name
    if payment.contact:
        return payment.contact.whatsapp_id
    return "Anonymous Giver"

# --- Member Export Functions ---

def export_members_to_excel(queryset):
    """
    Generates an Excel file with detailed information for the given queryset of MemberProfiles.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="member_details_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Member Details'

    headers = [
        "First Name", "Last Name", "WhatsApp ID", "Email", "Date of Birth",
        "Gender", "Marital Status", "Membership Status", "Date Joined",
        "City", "Country", "Notes"
    ]
    bold_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = bold_font

    for row_num, member in enumerate(queryset.select_related('contact'), 2):
        sheet.cell(row=row_num, column=1, value=member.first_name)
        sheet.cell(row=row_num, column=2, value=member.last_name)
        sheet.cell(row=row_num, column=3, value=member.contact.whatsapp_id if member.contact else "")
        sheet.cell(row=row_num, column=4, value=member.email)
        sheet.cell(row=row_num, column=5, value=member.date_of_birth)
        sheet.cell(row=row_num, column=6, value=member.get_gender_display())
        sheet.cell(row=row_num, column=7, value=member.get_marital_status_display())
        sheet.cell(row=row_num, column=8, value=member.get_membership_status_display())
        sheet.cell(row=row_num, column=9, value=member.date_joined)
        sheet.cell(row=row_num, column=10, value=member.city)
        sheet.cell(row=row_num, column=11, value=member.country)
        sheet.cell(row=row_num, column=12, value=member.notes)

    _auto_adjust_excel_columns(sheet)
    workbook.save(response)
    return response

def export_members_to_pdf(queryset):
    """
    Generates a PDF file with a summary of member details for the given queryset.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("Member Details Report", styles['h1']))
    elements.append(Spacer(1, 12))

    headers = ["Name", "WhatsApp ID", "Email", "Membership", "City", "Date Joined"]
    data = [headers]
    for member in queryset.select_related('contact'):
        data.append([
            member.get_full_name() or "",
            member.contact.whatsapp_id if member.contact else "",
            member.email or "",
            member.get_membership_status_display() or "",
            member.city or "",
            member.date_joined.strftime("%Y-%m-%d") if member.date_joined else ""
        ])

    table = Table(data, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F8B3A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F0F0')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="member_details_{timezone.now().strftime("%Y-%m-%d")}.pdf"'
    return response

# --- Payment Summary Export Functions ---

def export_payment_summary_to_excel(queryset, period_name):
    """
    Generates an Excel file summarizing payments by type for a given period.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="payment_summary_{period_name}_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Payment Summary - {period_name.title()}'

    bold_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    sheet.cell(row=1, column=1, value=f"Payment Summary for {period_name.replace('_', ' ').title()}").font = bold_font
    sheet.merge_cells('A1:C1')

    headers = ["Payment Type", "Total Amount (USD)", "Transaction Count"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_num, value=header)
        cell.font = header_font

    summary_data = queryset.values('payment_type').annotate(total_amount=Sum('amount'), transaction_count=Count('id')).order_by('payment_type')

    row_num = 4
    grand_total_amount = Decimal('0.00')
    grand_total_count = 0
    payment_type_display_map = dict(Payment.PAYMENT_TYPE_CHOICES)

    for summary in summary_data:
        payment_type_key = summary['payment_type']
        payment_type_display = payment_type_display_map.get(payment_type_key, payment_type_key.title())
        sheet.cell(row=row_num, column=1, value=payment_type_display)
        sheet.cell(row=row_num, column=2, value=summary['total_amount']).number_format = '"$"#,##0.00'
        sheet.cell(row=row_num, column=3, value=summary['transaction_count'])
        grand_total_amount += summary['total_amount']
        grand_total_count += summary['transaction_count']
        row_num += 1

    sheet.cell(row=row_num, column=1, value="Grand Total").font = header_font
    sheet.cell(row=row_num, column=2, value=grand_total_amount).font = header_font
    sheet.cell(row=row_num, column=2, value=grand_total_amount).number_format = '"$"#,##0.00'
    sheet.cell(row=row_num, column=3, value=grand_total_count).font = header_font

    _auto_adjust_excel_columns(sheet)
    workbook.save(response)
    return response

def export_payment_summary_to_pdf(queryset, period_name):
    """
    Generates a PDF file summarizing payments by type for a given period.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph(f"Payment Summary: {period_name.replace('_', ' ').title()}", styles['h1'])
    subtitle = Paragraph(f"Report generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal'])
    elements.extend([title, subtitle, Spacer(1, 24)])

    summary_data = queryset.values('payment_type').annotate(total_amount=Sum('amount'), transaction_count=Count('id')).order_by('payment_type')

    headers = ["Payment Type", "Total Amount (USD)", "Transactions"]
    data = [headers]
    grand_total_amount = Decimal('0.00')
    grand_total_count = 0
    payment_type_display_map = dict(Payment.PAYMENT_TYPE_CHOICES)

    for summary in summary_data:
        payment_type_key = summary['payment_type']
        total_amount = summary['total_amount']
        data.append([payment_type_display_map.get(payment_type_key, payment_type_key.title()), f"${total_amount:,.2f}", str(summary['transaction_count'])])
        grand_total_amount += total_amount
        grand_total_count += summary['transaction_count']

    data.append([Paragraph("<b>Grand Total</b>", styles['Normal']), Paragraph(f"<b>${grand_total_amount:,.2f}</b>", styles['Normal']), Paragraph(f"<b>{grand_total_count}</b>", styles['Normal'])])

    table = Table(data, colWidths=[200, 150, 100], hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F8B3A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#F0F0F0')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#C0C0C0')),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="payment_summary_{period_name}_{timezone.now().strftime("%Y-%m-%d")}.pdf"'
    return response

# --- Givers List Export Functions ---

def export_givers_list_finance_excel(queryset, period_name):
    """
    Generates an Excel file listing all givers and their total contributions for a period.
    For the finance department.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="givers_finance_report_{period_name}_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Givers Report (Finance)'

    bold_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    sheet.cell(row=1, column=1, value=f"Givers Report (Finance) for {period_name.replace('_', ' ').title()}").font = bold_font
    sheet.merge_cells('A1:B1')

    headers = ["Giver Name", "Total Amount (USD)"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_num, value=header)
        cell.font = header_font

    givers = {}
    for payment in queryset.select_related('member', 'contact'):
        contact_id = payment.contact_id
        if contact_id not in givers:
            givers[contact_id] = {'name': get_giver_name(payment), 'total': Decimal('0.00')}
        givers[contact_id]['total'] += payment.amount

    sorted_givers = sorted(givers.values(), key=lambda x: x['name'])
    
    row_num = 4
    for giver in sorted_givers:
        sheet.cell(row=row_num, column=1, value=giver['name'])
        sheet.cell(row=row_num, column=2, value=giver['total']).number_format = '"$"#,##0.00'
        row_num += 1

    _auto_adjust_excel_columns(sheet)
    workbook.save(response)
    return response

def export_givers_list_finance_pdf(queryset, period_name):
    """
    Generates a PDF file listing all givers and their total contributions for a period.
    For the finance department.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph(f"Givers Report (Finance): {period_name.replace('_', ' ').title()}", styles['h1'])
    elements.extend([title, Spacer(1, 24)])

    givers = {}
    for payment in queryset.select_related('member', 'contact'):
        contact_id = payment.contact_id
        if contact_id not in givers:
            givers[contact_id] = {'name': get_giver_name(payment), 'total': Decimal('0.00')}
        givers[contact_id]['total'] += payment.amount

    sorted_givers = sorted(givers.values(), key=lambda x: x['name'])

    headers = ["Giver Name", "Total Amount (USD)"]
    data = [headers]
    for giver in sorted_givers:
        data.append([giver['name'], f"${giver['total']:,.2f}"])

    table = Table(data, colWidths=[300, 150], hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F8B3A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F0F0')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="givers_finance_report_{period_name}_{timezone.now().strftime("%Y-%m-%d")}.pdf"'
    return response

def export_givers_list_publication_excel(queryset, period_name):
    """
    Generates an Excel file listing the names of all givers for public acknowledgment.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="givers_publication_list_{period_name}_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Givers List (Publication)'

    bold_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    sheet.cell(row=1, column=1, value=f"Thank You To Our Givers for {period_name.replace('_', ' ').title()}").font = bold_font
    sheet.merge_cells('A1:A1')

    sheet.cell(row=3, column=1, value="Giver Name").font = header_font

    giver_names = sorted(list(set(get_giver_name(p) for p in queryset.select_related('member', 'contact'))))

    for row_num, name in enumerate(giver_names, 4):
        sheet.cell(row=row_num, column=1, value=name)

    _auto_adjust_excel_columns(sheet)
    workbook.save(response)
    return response

def export_givers_list_publication_pdf(queryset, period_name):
    """
    Generates a PDF file listing the names of all givers for public acknowledgment.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph(f"A Special Thank You To Our Givers", styles['h1'])
    subtitle = Paragraph(f"For {period_name.replace('_', ' ').title()}", styles['h2'])
    elements.extend([title, subtitle, Spacer(1, 24)])

    giver_names = sorted(list(set(get_giver_name(p) for p in queryset.select_related('member', 'contact'))))

    data = [["Giver Name"]] + [[name] for name in giver_names]

    table = Table(data, colWidths=[450], hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F8B3A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F0F0')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="givers_publication_list_{period_name}_{timezone.now().strftime("%Y-%m-%d")}.pdf"'
    return response
